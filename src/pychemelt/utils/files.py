"""
This module contains helper functions to parse Differential Scanning Fluorimetry files from different instrument providers
Author: Osvaldo Burastero

All functions that import files should return:
- signal_data_dic: dictionary with the signal data, one entry per signal
- temp_data_dic: dictionary with the temperature data, one entry per signal
- conditions: list with the names of the samples
- signals: list with the names of the signals

A signal can be "350nm", "330nm", "Scattering", "Ratio", "Turbidity", "Ratio 350nm/330nm", etc.
The length of the lists in signal_data_dic and temp_data_dic should be the same as the length of conditions
"""

import numpy  as np
import pandas as pd

import codecs
import json
import re
import os

from openpyxl import load_workbook
from xlrd     import open_workbook

from collections import Counter

from .processing import *

__all__ = [
    "load_csv_file",
    "load_aunty_xlsx",
    "load_quantstudio_txt",
    "load_thermofluor_xlsx",
    "load_nanoDSF_xlsx",
    "load_panta_xlsx",
    "load_uncle_multi_channel",
    "load_mx3005p_txt",
    "detect_file_type",
    "detect_encoding",
    "read_jasco_thermal_ramp"
]

def get_sheet_names_of_xlsx(filepath):
    """
    Get the sheet names of a xls or xlsx file without loading it.
    The open_workbook function is used so we can handle the error "openpyxl does not support the old .xls file format"

    Parameters
    ----------
    filepath : str
        Path to the xls or xlsx file

    Returns
    -------
    list
        List of sheet names
    """

    try:

        wb = load_workbook(filepath, read_only=True, keep_links=False)
        sheetnames = wb.sheetnames

    except:

        xls = open_workbook(filepath, on_demand=True)
        sheetnames = xls.sheet_names()

    return sheetnames

def file_is_of_type_aunty(file_path):

    """
    Detect if file is an AUNTY xlsx file.

    The AUNTY format contains multiple sheets where the first column is
    temperatures and subsequent columns are fluorescence values. The first
    row contains the word 'wavelength' and the second row contains the word
    'temperature' in the first column.

    Parameters
    ----------
        file_path (str): Path to the .xls or .xlsx file to test.

    Returns
    -------
        bool: True if the file matches the AUNTY format heuristic, False otherwise.
    """

    if not (file_path.endswith('.xls') or file_path.endswith('.xlsx')):
        return False

    sheet_names = get_sheet_names_of_xlsx(file_path)

    for sheet_name in sheet_names:

        # Load the data
        data = pd.read_excel(
            file_path, sheet_name=sheet_name,
            header=None,skiprows=0)

        try:

            wavelength_cell = data.iloc[0, 1]
            temperature_cell = data.iloc[1, 0]
            corner_cell = data.iloc[0, 0]

            # Verify that the word wavelength is in the first row, second column
            condition1 = isinstance(wavelength_cell, str) and 'wavelength' in wavelength_cell.lower()

            # Verify that the word temperature is in the second row, first column
            condition2 = isinstance(temperature_cell, str) and 'temperature' in temperature_cell.lower()

            # Verify that the corner cell is empty (NaN)
            condition3 = np.isnan(corner_cell)

            if not (condition1 and condition2 and condition3):
                continue

            return True

        except:

            file_is_aunty = False

    return file_is_aunty

def file_is_of_type_uncle(xlsx_file):

    """
    Check if the file is an uncle file

    Parameters
    ----------
    xlsx_file : str
        Path to the xlsx file

    Returns
    -------
    bool
        True if the file is an uncle file, False otherwise
    """

    try:

        # Read the first sheet of the xlsx file
        data = pd.read_excel(xlsx_file,skiprows=1,nrows=5,header=None)

        # Extrac the first row
        row = data.iloc[0,1:]

        # concatenate the row into a string
        row_str = ' '.join([str(x) for x in row])

        # count the number of times we have word 'Time' and 'Temp'
        count_time = row_str.count('Time')
        count_temp = row_str.count('Temp')

        if count_time > 20 and count_temp > 20 and count_temp == count_time:
            return True
        else:
            return False
    except:
        return False

def detect_file_type(file):

    """
    Detect the type of file based on its extension and content.

    Parameters
    ----------
    file : str
        Path to the file

    Returns
    -------
    str or None
        Type of file (e.g., 'supr', 'csv', 'prometheus', 'panta', 'uncle', 'mx3005p', 'quantstudio', etc.) or None if unknown
    """

    file_extension = file.split('.')[-1]

    if file_extension == 'supr':
        return 'supr'

    if file_extension in ["xlsx", "xls"]:
        # Get file type: DSF or nDSF
        sheet_names = get_sheet_names_of_xlsx(file)

        # Load the data to the Python class
        if "RFU" in sheet_names:
            return 'thermofluor'
        elif "Data Export" in sheet_names or "melting-scan" in sheet_names:
            return 'panta'
        elif file_is_of_type_uncle(file):
            return 'uncle'
        elif file_is_of_type_aunty(file):
            return 'aunty'
        else:
            return 'prometheus'

    if file_is_jasco_thermal_ramp(file):
        return "jasco_thermal_ramp"

    if file_is_jasco_thermal_ramp_format_2(file):
        return "jasco_thermal_ramp_format_2"

    if file_extension == 'csv':

        if file_is_chirascan_thermal_ramp(file):

            return "chirascan_thermal_ramp"

        else:

            return 'csv'

    with codecs.open(file, 'r', encoding='utf-8',errors='ignore') as rf:
        ls       = rf.read().splitlines()

        for line in ls:
            if line.startswith('Segment') and 'Well' in line:
                return 'mx3005p'

            splittedLine = line.split()
            if 'Well' in splittedLine and 'Target' in splittedLine and 'Reading' in splittedLine:
                return 'quantstudio'

    # Return an error if no extension is found
    raise ValueError(f'File extension not recognized: {file_extension}')

def load_aunty_xlsx(file_path):

    """
    Load AUNTY-format multi-sheet Excel file where each sheet is a condition.

    Parameters
    ----------
    file_path : str
        Path to the AUNTY xlsx file
    """
    # Get the names of the sheets
    sheet_names = get_sheet_names_of_xlsx(file_path)

    wavelengths       = None
    temperatures      = []

    conditions        = []
    signals_data      = []

    signal_data_dic = {}
    temp_data_dic   = {}

    for sheet_name in sheet_names:

        # Load the data
        data = pd.read_excel(
            file_path, sheet_name=sheet_name,
            header=None,skiprows=0
            )

        wavelength_cell = data.iloc[0, 1]
        temperature_cell = data.iloc[1, 0]
        corner_cell = str(data.iloc[0, 0])

        # Verify that the word wavelength is in the first row, second column
        condition1 = isinstance(wavelength_cell, str) and 'wavelength' in wavelength_cell.lower()

        # Verify that the word temperature is in the second row, first column
        condition2 = isinstance(temperature_cell, str) and 'temperature' in temperature_cell.lower()
        
        # Verify that the corner cell is empty (NaN)
        condition3 = corner_cell == '' or corner_cell.lower() == 'nan'

        if not (condition1 and condition2 and condition3):
            continue

        signal_data  = np.array(data.iloc[2:, 1:]).astype(float)

        if wavelengths is None:

            wavelengths    = np.round(np.array(data.iloc[1, 1:]).astype(float),2)

        temperature_data   = np.round(np.array(data.iloc[2:, 0]).astype(float), 2)

        signals_data.append(signal_data)
        temperatures.append(temperature_data)
        conditions.append(sheet_name)


    # Now we have one signal matrix per condition, but we need one signal matrix per wavelength, with
    # the conditions as columns
    # Therefore, for each wavelength, we create one dataframe per condition and then merge them

    named_wls = [str(wl) + ' nm' for wl in wavelengths]

    for i, named_wl in enumerate(named_wls):

        signal_data_as_list = [arr[:,i] for arr in signals_data]

        signal_data_dic[named_wl] = signal_data_as_list
        temp_data_dic[named_wl]   = temperatures

    return signal_data_dic, temp_data_dic, conditions, named_wls   

def detect_encoding(file_path):

    """
    Detect the encoding of a file by trying common encodings.

    Parameters
    ----------
    file_path : str
        Path to the file

    Returns
    -------
    str
        Detected encoding or the string 'Unknown encoding'
    """

    try:
        with codecs.open(file_path, encoding="utf-8", errors="strict") as f:
            f.read()
        return 'utf-8'
    
    except:
        with codecs.open(file_path, encoding="latin1", errors="strict") as f:
            f.read()
        return 'latin1'

def find_indexes_of_non_signal_conditions(signal_data,conditions):

    """
    Given the signal data and the conditions, find the indexes of the conditions that are not signal data.
    We assume that conditions with the word 'derivative' or 'S.D.' in it are not signal data.

    Parameters
    ----------
    signal_data : numpy.ndarray
        2D array with the signal data (shape: temperature x conditions)
    conditions : list
        List of condition names

    Returns
    -------
    list
        List of indexes to remove
    """

    # Find the indexes of conditions with derivative in it, or 'S.D.' in it
    idx_to_remove1 = [i for i, cond in enumerate(conditions) if 'derivative' in cond.lower() or 's.d.' in cond.lower()]

    # Find indexes empty signal_data columns
    idx_to_remove2 = [i for i in range(signal_data.shape[1]) if np.isnan(signal_data[:, i]).all()]

    # Find columns where all values are NaN
    nan_columns = np.all(np.isnan(signal_data), axis=0)

    # Get the column indices
    idx_to_remove3 = np.where(nan_columns)[0].tolist()

    # Find unique indexes to remove
    idx_to_remove = list(set(idx_to_remove1 + idx_to_remove2 + idx_to_remove3))

    return idx_to_remove

def find_repeated_words(string_lst):

    """
    Given a list of strings, find the repeated words in the list.

    Parameters
    ----------
    string_lst : list of str
        List of strings to analyze

    Returns
    -------
    list
        List of repeated words
    """

    # Repeated words
    words = " ".join(string_lst).split()

    # Count the occurrences of each word
    word_counts = Counter(words)

    # Find words that are repeated (appear, at least, as many times as elements in string_lst)
    repeated_words = [word for word, count in word_counts.items() if count >= len(string_lst)]

    return repeated_words

def remove_words_in_string(input_string,word_list):

    """
    Given a string and a list of words, remove the words from the string.

    Parameters
    ----------
    input_string : str
        String to be processed
    word_list : list of str
        List of words to remove

    Returns
    -------
    str
        Processed string
    """

    # Split the string into words
    words = input_string.split()

    # Replace words from the list with an empty string
    filtered_words = [word for word in words if word not in word_list]

    # Join the words back into a single string
    output_string = ' '.join(filtered_words)

    return output_string

def load_csv_file(file):

    """
    Load a CSV file containing temperature and signal columns and return structured data.

    Parameters
    ----------
    file : str
        Path to the csv file

    Returns
    -------
    signal_data_dic : dict
        Dictionary mapping signal names to lists of 1D numpy arrays (one array per condition)
    temp_data_dic : dict
        Dictionary mapping signal names to lists of temperature arrays corresponding to the signals
    conditions : list
        List of condition names
    signals : numpy.ndarray
        Array of signal name strings
    """

    signal_data_dic = {}
    temp_data_dic   = {}

    signals = []

    encoding = detect_encoding(file)

    # Try common delimiters
    for delimiter in [',', ';', '\t']:

        try:

            dat = pd.read_csv(file, delimiter=delimiter, encoding=encoding)

            # Convert non-numeric columns to NaN
            dat = dat.apply(pd.to_numeric, errors='coerce')

            # Produce error if we don't have 2 or more columns
            if len(dat.columns) < 2:
                raise ValueError('File does not have enough columns')

            # Set the conditions names and start index for the signal data
            if 'time' in dat.columns[0].lower():
                conditions = [str(c) for c in dat.columns[2:]]
                idx_start = 1
            else:
                conditions = [str(c) for c in dat.columns[1:]]
                idx_start = 0

            signal_data = np.array(dat.iloc[:, (idx_start + 1):]).astype('float')
            temperature_data = np.array(dat.iloc[:, idx_start]).astype('float')

            break

        except:

            pass

    idx_to_remove = find_indexes_of_non_signal_conditions(signal_data, conditions)

    # Remove the elements from the conditions array
    conditions = [cond for i, cond in enumerate(conditions) if i not in idx_to_remove]

    # Remove the columns from the signal data
    signal_data = np.delete(signal_data, idx_to_remove, axis=1)

    # Divide the conditions into groups, according to the presence of the words '350', '330', 'ratio', 'scattering'
    conditions_350nm = [cond for cond in conditions if '350' in cond.lower() and 'ratio' not in cond.lower()]
    conditions_330nm = [cond for cond in conditions if '330' in cond.lower() and 'ratio' not in cond.lower()]
    conditions_ratio = [cond for cond in conditions if 'ratio' in cond.lower()]
    conditions_scatt = [cond for cond in conditions if 'scattering' in cond.lower()]

    # Check that the number of conditions is the same for all groups
    conditions_lst = [conditions_350nm, conditions_330nm, conditions_ratio, conditions_scatt]
    sel_cond_lst   = [cond for cond in conditions_lst if len(cond) > 0]
    n_conditions   = [len(cond) for cond in conditions_lst if len(cond) > 0]

    possible_rep_conditions = len(np.unique(n_conditions)) == 1

    # case where the CSV corresponds to a nDSF csv
    if possible_rep_conditions:

        # For each group, store the signal data and the temperature data
        if len(conditions_350nm) > 0:
            signal_350 = signal_data[:, [i for i, cond in enumerate(conditions) if cond in conditions_350nm]]
            signal_data_as_list = signal_350.transpose().tolist()
            signal_data_dic["350 nm"] = signal_data_as_list
            temp_data_dic["350 nm"] = expand_temperature_list([temperature_data],signal_data_as_list)
            signals.append("350 nm")

        if len(conditions_330nm) > 0:
            signal_330 = signal_data[:, [i for i, cond in enumerate(conditions) if cond in conditions_330nm]]
            signal_data_as_list = signal_330.transpose().tolist()
            signal_data_dic["330 nm"] = signal_data_as_list
            temp_data_dic["330 nm"] = expand_temperature_list([temperature_data],signal_data_as_list)
            signals.append("330 nm")

        if len(conditions_ratio) > 0:
            signal_ratio = signal_data[:, [i for i, cond in enumerate(conditions) if cond in conditions_ratio]]
            signal_data_as_list = signal_ratio.transpose().tolist()
            signal_data_dic["Ratio 350 nm / 330 nm"] = signal_data_as_list
            temp_data_dic["Ratio 350 nm / 330 nm"] = expand_temperature_list([temperature_data],signal_data_as_list)
            signals.append("Ratio 350 nm / 330 nm")

        if len(conditions_scatt) > 0:
            signal_scatt = signal_data[:, [i for i, cond in enumerate(conditions) if cond in conditions_scatt]]
            signal_data_as_list = signal_scatt.transpose().tolist()
            signal_data_dic["Scattering"] = signal_data_as_list
            temp_data_dic["Scattering"] = expand_temperature_list([temperature_data],signal_data_as_list)
            signals.append("Scattering")

        cond_temp = sel_cond_lst[0]

        if len(cond_temp) > 1:

            repeated_words = find_repeated_words(cond_temp)
            cond_temp = [remove_words_in_string(cond, repeated_words) for cond in cond_temp]
            conditions = cond_temp

        else:

            conditions = cond_temp

    else:

        # If we only have one condition, use the condition as signal name
        if len(conditions) == 1:
            signal_name = conditions[0]
        # Default signal name
        else:
            signal_name = "Fluorescence"

        signal_data_as_list = signal_data.transpose().tolist()
        signal_data_dic[signal_name] = signal_data_as_list
        temp_data_dic[signal_name] = expand_temperature_list([temperature_data],signal_data_as_list)
        signals.append(signal_name)

    signals = np.array(signals)

    return signal_data_dic, temp_data_dic, conditions, signals

def get_start_line_quantstudio_txt(file_name):

    """
    Find the start line for QuantStudio text files (first non-comment data line).

    Parameters
    ----------
    file_name : str
        Path to the QuantStudio txt file

    Returns
    -------
    int
        Number  of the first line not starting with '*' plus one
    """

    with codecs.open(file_name, 'r', encoding='utf-8',errors='ignore') as rf:
        ls       = rf.read().splitlines()
        splitted = [l.split() for l in ls]
        
        for i,s in enumerate(splitted):
            if len(s) > 5 and "*" not in s[0]:
                start_row = i+1
                break

    return(start_row)

def load_quantstudio_txt(QSfile):

    """
    Load QuantStudio TXT files (.txt) exported from QuantStudio instruments.

    Parameters
    ----------
    QSfile : str
        Path to the QuantStudio txt file

    Returns
    -------
    signal_data_dic : dict
        Dictionary with signal data (key: 'Fluorescence')
    temp_data_dic : dict
        Dictionary with temperature arrays per condition
    conditions : list
        List of condition names (well identifiers)
    signals : numpy.ndarray
        Array with signal name(s)
    """

    start_row = get_start_line_quantstudio_txt(QSfile)
    data      = pd.read_csv(QSfile,skiprows=start_row,sep=r"\s+",header=None)

    signal_data_dic = {'Fluorescence': []}
    temp_data_dic   = {'Fluorescence': []}

    u, ind     = np.unique(data.iloc[:,1], return_index=True)
    conditions = u[np.argsort(ind)].tolist()

    signal = "Fluorescence"

    # Process each group directly without creating DataFrames
    for _, group in data.groupby(1):
        
        # Extract temperature and signal columns (columns 3 and 4, 0-indexed)
        temp_data = group.iloc[:, 3].astype(str).str.replace(',', '').astype(float).values
        signal_data = group.iloc[:, 4].astype(str).str.replace(',', '').astype(float).values
        
        # Sort by temperature
        sorted_indices = np.argsort(temp_data)
        temp_data = temp_data[sorted_indices]
        signal_data = signal_data[sorted_indices]

        signal_data_dic[signal].append(signal_data)
        temp_data_dic[signal].append(temp_data)

    signals = np.array([signal])

    return signal_data_dic, temp_data_dic, conditions, signals

def load_thermofluor_xlsx(thermofluor_file):

    """
    Load DSF Thermofluor xls file and extract data.

    Parameters
    ----------
    thermofluor_file : str
        Path to the xls file

    Returns
    -------
    signal_data_dic : dict
        Dictionary with signal data
    temp_data_dic : dict
        Dictionary with temperature data
    conditions : list
        List of conditions
    """

    xls  = pd.ExcelFile(thermofluor_file)
    dat = pd.read_excel(xls, "RFU",header=None)
    conditions = np.array(dat.iloc[0, 1:]).tolist()

    signal_data_dic = {}
    temp_data_dic   = {}

    signal = "DSF_RFU"

    fluo = np.array(dat.iloc[1:,1:]).astype('float').transpose().tolist()
    temp = [np.array(dat.iloc[1:, 0]).astype('float')]

    signal_data_dic[signal] = fluo
    temp_data_dic[signal] = expand_temperature_list(temp, fluo)

    signals = np.array([signal],dtype='str')

    return signal_data_dic, temp_data_dic, conditions, signals

def space_combinations(word):
    """
    Generate all combinations of a word with 0, 1, or 2 spaces before and/or after the word.

    Parameters
    ----------
    word : str
        The word to generate combinations for.

    Returns
    -------
    list
        List of strings with all combinations of spaces before and after the word.
    """
    spaces = [' ', '  ', '   ']
    combos = []
    for before in ['', *spaces]:
        for after in ['', *spaces]:
            combos.append(f"{before}{word}{after}")
    return combos

def load_nanoDSF_xlsx(processed_dsf_file):

    """
    Load nanotemper processed xlsx file and extract relevant data.

    Parameters
    ----------
    processed_dsf_file : str
        Path to the processed xlsx file

    Returns
    -------
    signal_data_dic : dict
        Dictionary with signal data
    temp_data_dic : dict
        Dictionary with temperature data
    conditions : list
        List of conditions
    signals : numpy.ndarray
        Array of signal names
    """

    sheet_names = get_sheet_names_of_xlsx(processed_dsf_file)

    signal_data_dic = {}
    temp_data_dic   = {}

    # this needs to be the processed file!
    xls = pd.ExcelFile(processed_dsf_file)

    conditions_df = pd.read_excel(xls, "Overview")
    conditions = conditions_df['Sample ID'].fillna('').astype(str).tolist()

    # Add position index to avoid problems with empty names

    possible_signals = space_combinations('350nm') + space_combinations('330nm') + space_combinations('Ratio') + space_combinations('Scattering')

    include = []

    # Change signal name if unfolding curve is present
    for sn in sheet_names:
        include_value = any(
            [ps in sn and "deriv" not in sn.lower() and "fold" not in sn.lower() for ps in possible_signals])

        include.append(include_value)

    sheet_names_to_load = np.array([s for (i, s) in zip(include, sheet_names) if i])

    signals = np.array([" ".join(sn.split()) for sn in sheet_names_to_load])

    for sn, signal in zip(sheet_names_to_load, signals):

        dat = pd.read_excel(xls, sn, index_col=None, header=None)

        indices   = np.argwhere(dat.iloc[:, 0].values == 'Time [s]')
        first_row = int(indices[0][0]) + 1

        # Find all the columns with the word temperature
        column_headers = dat.iloc[first_row - 1, :].values

        ids_temperature = [i for i, x in enumerate(column_headers) if 'temperature' in x.lower()]
        ids_time        = [i for i, x in enumerate(column_headers) if 'time'        in x.lower()]

        # If we have more than one temperature column, remove the extra temperature and time columns
        if len(ids_temperature) > 1:
            ids_temperature = ids_temperature[1:]  # We keep the first temperature colummn
            ids_time        = ids_time[1:]  # We keep the first time column

            # Remove columns of dataframe by index
            dat = dat.drop(dat.columns[ids_temperature + ids_time], axis=1)

        fluo = np.array(dat.iloc[first_row:, 2:]).astype('float').transpose().tolist()
        temp = [np.array(dat.iloc[first_row:, 1]).astype('float')]

        signal_data_dic[signal] = fluo
        temp_data_dic[signal]   = expand_temperature_list(temp,fluo)

    return signal_data_dic, temp_data_dic, conditions, signals

def load_panta_xlsx(pantaFile):

    """
    Load the xlsx file generated by a Prometheus Panta instrument.

    Parameters
    ----------
    pantaFile : str
        Path to the xlsx file

    Returns
    -------
    signal_data_dic : dict
        Dictionary with signal data
    temp_data_dic : dict
        Dictionary with temperature data
    conditions : list
        List of conditions
    signals : numpy.ndarray
        List of signal names, such as 330nm and 350nm
    """

    sheet_names = get_sheet_names_of_xlsx(pantaFile)

    data_sheetname = "Data Export" if "Data Export" in sheet_names else "melting-scan"

    data = pd.read_excel(pantaFile, data_sheetname)

    signal_data_dic = {}
    temp_data_dic   = {}

    column_names  = [str.lower(c) for c in data.columns]

    pos_350       = [i for i,x in enumerate(column_names) if "350"        in x and "deriv" not in x and "330" not in x]
    pos_330       = [i for i,x in enumerate(column_names) if "330"        in x and "deriv" not in x and "350" not in x]
    scattering    = [i for i,x in enumerate(column_names) if "scattering" in x and "deriv" not in x]
    pos_ratio     = [i for i,x in enumerate(column_names) if "ratio"      in x and "deriv" not in x]
    pos_turb      = [i for i,x in enumerate(column_names) if "turbidity"  in x and "deriv" not in x]

    possible_signals    = ["350 nm","330 nm","Scattering","Ratio","Turbidity"]
    signals             = []

    all_positions = [pos_350,pos_330,scattering,pos_ratio,pos_turb]

    for positions,signal in zip(all_positions,possible_signals):

        signal_data = []
        temp_data   = []

        if len(positions) > 0:

            for position in positions:

                # Find the index of the previous column with the word temperature
                temp_col = [i for i, x in enumerate(column_names) if 'temperature' in x.lower() and i < position][-1]

                # If the temperature column has the word "refolding", exclude it
                if "refolding" in column_names[temp_col]:
                    continue

                # round to two digits the temperature values
                temperature_values = data.iloc[:, temp_col].to_numpy().astype('float')
                temperature_values = np.round(temperature_values, 3)

                signal_values = data.iloc[:, position].to_numpy().astype('float')

                # Sort the temperature values and signal values together, according to the temperature values
                sorted_indices     = np.argsort(temperature_values)
                temperature_values = temperature_values[sorted_indices]
                signal_values      = signal_values[sorted_indices]

                # Extract the temperature data and signal data
                temp_data.append(temperature_values)
                signal_data.append(signal_values)

            signal_data_dic[signal]   = signal_data
            temp_data_dic[signal]     = temp_data

            signals.append(signal)

    signals        = np.array(signals)

    try:

        conditions_df = pd.read_excel(pantaFile, "Overview")
        conditions    = conditions_df[['Sample ID']].values.flatten().tolist()

    except:

        conditions = np.repeat('',len(signal_data)).tolist()

    return signal_data_dic, temp_data_dic, conditions, signals

def load_uncle_multi_channel(uncle_file):

    """
    Function to load the data from the UNCLE instrument.

    Parameters
    ----------
    uncle_file : str
        Path to the xlsx file

    Returns
    -------
    signal_data_dic : dict
        Dictionary with signal data (keys: wavelength strings like '350 nm')
    temp_data_dic : dict
        Dictionary with temperature arrays per condition
    conditions : list
        List of sample names
    signals : list
        List of wavelength strings
    """

    # Get the names of the sheets
    sheet_names = get_sheet_names_of_xlsx(uncle_file)

    wavelengths       = None
    temperatures      = []

    conditions        = []
    signals_data      = []

    signal_data_dic = {}
    temp_data_dic   = {}

    # Loop through each sheet and read the data
    for sheet_name in sheet_names:

        try:

            # Read the data from the sheet
            data = pd.read_excel(
                uncle_file, 
                sheet_name=sheet_name,
                header=None,
                skiprows=0
                )

            # Extract the sample name, from the first row, fitfh column
            sample_name = data.iloc[0, 4]

            # Remove the first row
            data = data.iloc[1:, :]

            # Extract the time/temperature data
            temperature_data = data.iloc[0, 1:].values

            # Select the temperature data
            temperature_data = [x.split(',')[0] for x in temperature_data]
            temperature_data = [x.split(':')[1] for x in temperature_data]
            temperature_data = np.array(temperature_data,dtype=float)

            # Extract the signal data
            # It contains one column per temperature and one row per wavelength
            signal_data = np.array(
                data.iloc[3:, 1:].values,
                dtype=float
                )

            # Assign the wavelength data if wavelengths is None
            if wavelengths is None:
                wavelengths = np.round(
                    np.array(
                        data.iloc[3:, 0].values,
                        dtype=float
                        ),
                    decimals=1
                    )

            signals_data.append(signal_data) # List of 2D arrays, one per condition
            temperatures.append(temperature_data) # List of 1D arrays, one per condition
            conditions.append(sample_name)

        except:

            pass

    # Iterate over the wavelengths
    named_wls = [str(wl) + ' nm' for wl in wavelengths]

    # We require one signal matrix per wavelength
    # with one column per condition
    for i in range(len(wavelengths)):

        # Extract the signal data for the current wavelength
        # signal_temp has one row per condition and one column per wavelength
        signal_data_i = [arr[i,:] for arr in signals_data]

        wl = named_wls[i]

        signal_data_dic[wl] = signal_data_i
        temp_data_dic[wl]   = temperatures

    return signal_data_dic, temp_data_dic, conditions, named_wls

def load_mx3005p_txt(filename):
    """
    Load Agilent MX3005P qPCR txt file and extract data

    Parameters
    ----------
    filename : str
        Path to the MX3005P txt file. The second column has the fluorescence data, and
        the third column the temperature. Wells are separated by rows containing a sentence like this one: 'Segment  2 Plateau  1 Well  1'

    Returns
    -------
    signal_data_dic : dict
        Dictionary with signal data
    temp_data_dic : dict
        Dictionary with temperature data
    conditions : list
        List of conditions (well numbers)
    signals : numpy.ndarray
        List of signal names
    """
    
    signal_data_dic = {'Fluorescence': []}
    temp_data_dic = {'Fluorescence': []}
    conditions = []
    
    with open(filename, 'r') as f:
        ls = f.read().splitlines()

        for i, line in enumerate(ls):
            if line.startswith('Segment') and 'Well' in line:
                # Get the well number
                well_num = line.split()[-1]
                conditions.append(well_num)

                fluorescence = []
                temperature = []

                # Read data for this well
                for line2 in ls[i+2:]:
                    if line2.startswith('Segment'):
                        break
                    else:
                        data = line2.split()
                        if len(data) >= 3:
                            fluorescence.append(float(data[1]))
                            temperature.append(float(data[2]))

                # Convert to arrays and sort by temperature
                temp_array = np.array(temperature)
                fluo_array = np.array(fluorescence)
                
                sorted_indices = np.argsort(temp_array)
                temp_array = temp_array[sorted_indices]
                fluo_array = fluo_array[sorted_indices]
                
                signal_data_dic['Fluorescence'].append(fluo_array)
                temp_data_dic['Fluorescence'].append(temp_array)

    signals = np.array(['Fluorescence'])
    
    return signal_data_dic, temp_data_dic, conditions, signals

def load_supr_dsf( JSON_file):

    signal_data_dic = {}
    temp_data_dic   = {}

    # Read JSON data from a file
    with open(JSON_file, "r") as file:
        json_data = file.read()

    # Parse JSON data into a dictionary
    data_dict = json.loads(json_data)

    samples_name = [item["SampleName"] for item in data_dict['Samples']]
    samples_well = [item["WellLocations"] for item in data_dict['Samples']]

    samples_name_simple = []
    samples_well_simple = []

    for sn, sw in zip(samples_name, samples_well):

        if ',' in sw:

            sw = sw.split(',')
            sn = [sn for _ in sw]

        else:

            sw = [sw]
            sn = [sn]

        samples_name_simple += sn
        samples_well_simple += sw

    name_df = pd.DataFrame({
        'well': samples_well_simple,
        'name': samples_name_simple})

    scans = [item["_scans"] for item in data_dict['Wells']]
    n_scans = len(scans)

    wavelengths = data_dict['Wavelengths']
    wavelengths = np.round(wavelengths, decimals=1)
    n_wavelengths = len(wavelengths)

    temperatures = []
    signals = []

    well = [item["PhysicalLocation"] for item in data_dict['Wells']]

    # Create a categorical data type with the custom order
    cat_type = pd.CategoricalDtype(categories=well, ordered=True)

    # Convert the column to the categorical data type
    name_df['well'] = name_df['well'].astype(cat_type)

    # Sort the DataFrame based on the custom order
    name_df = name_df.sort_values(by='well')

    conditions = name_df['name'].values.astype(str).tolist()

    for i in range(n_scans):
        temperatures.append([item['Temperature'] for item in scans[i]])
        signals.append([item['Signal'] for item in scans[i]])

    temperatures = np.array(temperatures).T
    temperatures = np.round(temperatures, decimals=1)

    signals = np.array(signals)

    # Iterate over the wavelengths

    named_wls = [str(wl) + ' nm' for wl in wavelengths]

    for i in range(n_wavelengths):

        signals_temp = signals[:, :, i].T

        signals_data = []
        temps_data   = []

        # Iterate over the columns of the arrays
        for ii in range(n_scans):

            x = temperatures[:, ii]
            y = signals_temp[:, ii]

            sorted_indices = np.argsort(x)
            x = x[sorted_indices]
            y = y[sorted_indices]

            # Find np.nans in y and x, and remove them
            non_nas_y = np.logical_not(np.isnan(y))
            non_nas_x = np.logical_not(np.isnan(x))

            non_nas_yx = np.logical_and(non_nas_y, non_nas_x)

            x = x[non_nas_yx]
            y = y[non_nas_yx]

            signals_data.append(y)
            temps_data.append(x)

        wl = named_wls[i]

        signal_data_dic[wl] = signals_data
        temp_data_dic[wl] = temps_data

    return signal_data_dic, temp_data_dic, conditions, named_wls

def find_delimiter_character(string):

    """
    Guess if a string contains ',' separated values or ';' separated values

    Parameters
    ----------
    string : str
        The string to analyze

    Returns
    -------
    delimiter : str
        The delimiter character that is most likely used in the string (',' or ';' or '\t')
    """

    comma_count     = string.count(",")
    semicolon_count = string.count(";")
    tab_count       = string.count("\t")

    counts = {
        ",": comma_count,
        ";": semicolon_count,
        "\t": tab_count
    }

    delimiter = max(counts, key=counts.get)

    # Choose the delimiter with the highest count
    return delimiter


def read_jasco_thermal_ramp(file):

    """
    Given a JASCO file with a thermal ramp, this function reads the data

    The data is given in chuncks:

    Channel 1
	4.94	14.93	25.08	35.02	45.03	55.07	64.99	75.04	85.08	95.03
    250	-0.310564	-0.112003	0.0199744	-0.217282	-0.238716	-0.173046	0.00129784	-0.394731	-0.687165	-1.40543

    Parameters
    ----------
    file : str
        Path to the JASCO thermal ramp file

    Returns
    -------
    signal_data_dic : dict
        Dictionary with signal data (key: wavelength string like '250 nm')
    temp_data_dic : dict
        Dictionary with temperature arrays per condition (only one condition in this case)
    conditions : list
        File name as condition
    wavelength_data : list
        List of wavelength strings (e.g., '250 nm', '255 nm', etc.)

    """

    with open(file, encoding="latin-1") as f:

        ls = f.read().splitlines()

        split_str = find_delimiter_character("".join(ls[:20]))

        channels_idx =[i+15 for i, l in enumerate(ls[15:]) if "Channel" in l]

        # Find the first line with the word 'Channel'
        channel_idx1 = channels_idx[0]

        channel_idx2 = channels_idx[1] if len(channels_idx) > 1 else len(ls)

        # Find the number of temperature points (Line after channel 1)
        temperatures = ls[channel_idx1 + 1].split(split_str)

        # Remove empty lines
        temperatures = [parse_number(x) for x in temperatures if x != '']

        # Convert temperatures to numpy array
        temperatures = np.array(temperatures)

        # Find the wavelength data, first column
        wavelength_data = [str(parse_number(x.split(split_str)[0])) + " nm" for x in ls[(channel_idx1+2):channel_idx2]]

        # Convert wavelength_data to  numpy array
        wavelength_data = np.array(wavelength_data)

        # Retrieve the CD data
        cd_data = [x.split(split_str)[1:] for x in ls[(channel_idx1+2):channel_idx2]]

        for i in range (len(cd_data)):

            cd_data[i] = [parse_number(x) for x in cd_data[i] if x != '']

        signal_data_dic ={}
        temp_data_dic  = {}

        for cd_signal,wl in zip(cd_data,wavelength_data):

            signal_data_dic[wl] = [cd_signal]
            temp_data_dic[wl] = [temperatures]

        # extract file basename
        bname = os.path.basename(file)

        conditions = [bname]

    return signal_data_dic, temp_data_dic, conditions, wavelength_data


def file_is_jasco_thermal_ramp(file):

    """
    Find if a file is of type JASCO thermal ramp

    Parameters
    ----------
    file : str
        Path to the file

    Returns
    -------
    bool
        True if the file is a JASCO thermal ramp, False otherwise

    """

    with open(file, encoding="latin-1") as f:

        ls = f.read().splitlines()

        split_str = find_delimiter_character("".join(ls[:20]))

        condition0 = False

        # Detect if we have a thermal ramp by finding
        # the presence of many numeric columns
        for l in ls[20:50]:

            l_split     = l.split(split_str)

            if are_all_strings_numeric(l_split) and len(l_split) > 4:
                condition0 = True
                break

        l_split = [l.split(split_str) for l in ls[:20]]

        condition1 = any([x[0].lower() == "origin" and x[1].lower() == "jasco"      for x in l_split])
        condition2 = any([x[0].lower() == "xunits" and x[1].lower() == "nanometers" for x in l_split])

        if all([condition0, condition1, condition2]):
            return True

    return False

def file_is_jasco_thermal_ramp_format_2(file):

    """
    Find if a file is of type JASCO thermal ramp, format 2

    Parameters
    ----------
    file : str
        Path to the file

    Returns
    -------
    bool
        True if the file is a JASCO thermal ramp, format 2, False otherwise

    """

    with open(file, encoding="latin-1") as f:

        ls = f.read().splitlines()

        split_str = find_delimiter_character("".join(ls[:20]))

        condition0 = False

        # Detect if we have a thermal ramp by finding
        # the presence of numeric columns, but less than 10
        for l in ls[20:50]:

            l_split     = l.split(split_str)

            if are_all_strings_numeric(l_split) and len(l_split) >= 2 and len(l_split) < 10:
                condition0 = True
                break

        l_split = [l.split(split_str) for l in ls[:20]]

        condition1 = any([x[0].lower() == "origin" and "jasco" in x[1].lower()       for x in l_split])
        condition2 = any([x[0].lower() == "xunits" and "temperature" in x[1].lower() for x in l_split])

        if all([condition0, condition1, condition2]):
            return True

    return False

def read_jasco_thermal_ramp_format_2(file):

    """
    Given a JASCO file with a thermal ramp at only one wavelength, this function reads the data

    XYDATA
    4,0000	-29,1176	295,223
    6,0600	-29,3217	295,577
    8,0600	-29,0102	295,484
    10,0500	-28,9184	295,437
    12,0600	-28,8087	295,167
    14,1000	-28,5928	295,186


    """

    with open(file, encoding="latin-1") as f:

        ls = f.read().splitlines()

        split_str = find_delimiter_character("".join(ls[:20]))

        # Find the wavelength
        line_with_wl_info = [ l for l in ls if "wavelength" in l ][0]

        # Remove all non-numeric characters
        wavelength_value = re.sub("[^0-9,.-]", "", line_with_wl_info)

        # Find the first line with the word 'XYDATA'
        idx_start = [ i for i, l in enumerate(ls) if "XYDATA" in l ][0] + 1

        # Find the first empty line or starting with asterisk after idx_start
        idx_end = [i for i,l in enumerate(ls) if l == '' or l.startswith('*') if i > idx_start][0]

        sel_lines = ls[idx_start:idx_end]

        sel_lines_split = [ l.split(split_str) for l in sel_lines ]

        temperatures = [parse_number(l[0]) for l in sel_lines_split]
        temperatures = np.array(temperatures)

        cd_data = [parse_number(l[1]) for l in sel_lines_split]

    signals = np.array([str(parse_number(wavelength_value)) + " nm"])

    # extract file basename
    bname = os.path.basename(file)

    conditions = [bname]

    signal_data_dic = {signals[0]: [cd_data]}
    temp_data_dic   = {signals[0]: [temperatures]}

    return signal_data_dic, temp_data_dic, conditions, signals

def file_is_chirascan_thermal_ramp(cd_file):

    """
    Find if a file is of type Chirascan thermal ramp.

    Parameters
    ----------
    cd_file : str
        Path to the file

    Returns
    -------
    bool
        True if the file is a Chirascan thermal ramp, False otherwise

    """

    words_to_find = ["circulardichroism", "wavelength", "temperature"]

    with open(cd_file) as f:
        ls = f.read().splitlines()[:100]
        # Join the lines into a single string and convert to lowercase for easier searching
        text = " ".join(ls).lower()

        # Check if all the words are present in the text
        if all(word in text for word in words_to_find):
            return True

    return False

def read_chirascan_thermal_ramp(cd_file):

    """
    Read the CD data from a Chirascan file with a thermal ramp.

    Parameters
    ----------
    cd_file : str
        Path to the Chirascan file.

    Returns
    -------

    """

    # Open the file for reading
    with open(cd_file) as f:
        # Read all lines and split them into a list of lines
        ls = f.read().splitlines()

        # Find the row index where the numeric data starts (CD spectra)
        for i, l in enumerate(ls):
            # Check if the line contains "Wavelength"
            if "wavelength" in l.lower():
                # Assuming metadata starts two lines below the line containing "Wavelength"
                # Get the third row (i+2) and split it by comma to count the number of columns
                row2 = ls[i + 2].split(",")
                # Check if the first element of row2 is numeric, which indicates the start of data
                if row2[0].isnumeric():
                    # n_cols indicates the number of columns in the data
                    n_cols = len([x for x in row2 if x])
                    break

        # Initialize a list to store CD data, one list per row (wavelength)
        cd_data = []
        # Initialize an empty list to store wavelengths
        wavelength = []

        # Extract temperature values from the line after the "Wavelength" line
        temperature = ls[i + 1].split(",")[1:]
        temperature = [t for t in temperature if t]
        # Convert temperature values to a numpy array
        temperature = np.array(temperature).astype(float)

        # Iterate through the lines containing CD data (starting from row i+2)
        for i2, l in enumerate(ls[i + 2 :]):
            # Split the line using comma as the delimiter and remove any empty elements
            row = l.split(",")
            row = [x for x in row if x]

            # Check if the row has the correct number of columns (n_cols)
            if len(row) != n_cols:
                break

            # Append the wavelength value to the list
            wavelength.append(float(row[0]))

            # Convert
            signal_at_wavelength = np.array([parse_number(x) for x in row[1:]])

            cd_data.append(signal_at_wavelength)

        # Convert the list of wavelengths to a numpy array
        wavelength_data = np.array([str(wl) + " nm" for wl in wavelength])

        signal_data_dic ={}
        temp_data_dic  = {}

        for cd_signal,wl in zip(cd_data,wavelength_data):

            signal_data_dic[wl] = [cd_signal]
            temp_data_dic[wl] = [temperature]

        # extract file basename
        bname = os.path.basename(cd_file)

        conditions = [bname]

    return signal_data_dic, temp_data_dic, conditions, wavelength_data